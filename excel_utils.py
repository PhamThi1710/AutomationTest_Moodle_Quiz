import openpyxl
from openpyxl.styles import PatternFill

class excel_utils:
    def __init__(self, file, sheetname):
        self.file = file
        self.sheetname = sheetname
        self.workload = openpyxl.load_workbook(file)
        self.sheet = self.workload[sheetname]
    def get_row_count(self):
        return (self.sheet.max_row)
    def get_column_count(self):
        return (self.sheet.max_column)
    def read_data(self, row, col):
        return (self.sheet.cell(row, col).value)
    def write_data(self, row, col, data):
        self.sheet.cell(row, col).value = data
        self.workload.save(self.file) 
    def fill_Color(self, row, col, color):
        self.sheet.cell(row,col).fill = PatternFill(start_color=color, end_color=color,
                                        fill_type = "solid")
        self.workload.save(self.file) 
           